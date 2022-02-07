# Install the openpyxl library
from openpyxl import load_workbook
from yattag import Doc, indent, EACH_LINE

# for chdir()
import os

# 작업 디렉토리로 이동한다.
os.chdir("C:/workpy/ExcelToXML")

# Loading our Excel file
wb = load_workbook("./SRC/접속표준서(정보분배-UDP_TCP실시간)_v1.02.xlsx")

# Getting an object of active sheet 4th
ws = wb.worksheets[3]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()
  
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
  
# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)
  
with tag('InterfaceDefDoc'):    #인터페이스정의서
    for row in ws.iter_rows(min_row=3, max_row=3977, min_col=1, max_col=13):
        col = [cell.value for cell in row]
        print(col)
        with tag("Row"):
            with tag("Interface_ID"):
                text(col[0])
            with tag("Inferface_Name"):
                text(col[1])
            with tag("Sequence"):
                text(col[3])
            with tag("ItemName"):
                text(col[4])
            with tag("Depth"):
                text(col[6])
            with tag("DataType"):
                text(col[7])
            with tag("Length"):
                text(col[8])  
            with tag("AccLength"):
                text(col[9])
            with tag("Definition"):
                text(col[10])
            with tag("Remarks"):
                if col[12]:
                    text(col[12])

result = indent(
    doc.getvalue(),
    indentation='   ',
    newline = '\n',
    indent_text=False
)
  
with open("./XML/krx_spec_inf_splt.xml", "w", encoding='UTF-8') as f:
    f.write(result)
