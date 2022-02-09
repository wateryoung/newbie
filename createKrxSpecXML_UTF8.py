# Install the openpyxl library
from openpyxl import load_workbook
from yattag import Doc, indent

# for chdir()
import os

# 작업 디렉토리로 이동한다.
os.chdir("C:/workpy/ExcelToXML")

# Loading our Excel file
wb = load_workbook("./SRC/접속표준서(정보분배-UDP_TCP실시간)_v1.02.xlsx")

# Getting an object of active sheet 4th
ws = wb.worksheets[3]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()
  
xml_header = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
  
# Appends the String to document
doc.asis(xml_header)
#doc.asis(xml_schema)
  
with tag('InterfaceDefDoc'):    #인터페이스정의서
    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=13):
        col = [cell.value for cell in row]
#        print(col)
        with tag("Row"):
            with tag("InfcID"):         #인터페이스ID
                text(col[0])
            with tag("InfcNm"):         #인터페이스명
                text(col[1])
            with tag("Sno"):            #순번
                text(col[3])
            with tag("ItmNm"):          #항목명
                text(col[4])
            with tag("Dcpt"):           #소수점 이하 자릿수(Decimal Point)
                text(col[6])
            with tag("DataTp"):         #Data Type
                text(col[7])
            with tag("Lngh"):           #길이/
                text(col[8])
            with tag("AcltLngh"):       #누적길이
                text(col[9])
            with tag("Dfnt"):           #정의
                text(col[10])
            with tag("Rmks"):           #비고
                if col[12]:
                    text(col[12])

result = indent(
    doc.getvalue(),
    indentation='   ',
    newline = '\n',
    indent_text=False
)
  
with open("./XML/krx_spec_inf_splt_utf8.xml", "w", encoding='UTF-8') as f:
    f.write(result)
